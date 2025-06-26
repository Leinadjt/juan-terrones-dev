from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date, time
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
import io
import json
import openpyxl
from sqlalchemy import func, inspect, text
import pytz
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from collections import defaultdict

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///polleria.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# FILTRO JINJA PERSONALIZADO PARA HORA DE PERÚ
@app.template_filter('peru_time')
def _jinja2_filter_peru_time(utc_dt):
    if not utc_dt:
        return ""
    # Asume que utc_dt es un objeto datetime naive en UTC
    utc_dt = utc_dt.replace(tzinfo=pytz.utc)
    peru_tz = pytz.timezone('America/Lima')
    peru_dt = utc_dt.astimezone(peru_tz)
    return peru_dt.strftime('%H:%M')

# Modelos de la base de datos
class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    rol = db.Column(db.String(20), nullable=False)  # admin, cajero, cocinero
    activo = db.Column(db.Boolean, default=True)

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    precio = db.Column(db.Float, nullable=False)
    categoria = db.Column(db.String(50), nullable=False)  # combo, plato, bebida, postre
    activo = db.Column(db.Boolean, default=True)

class Insumo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    unidad = db.Column(db.String(20), nullable=False)  # kg, unidades, litros
    stock_actual = db.Column(db.Float, default=0)
    stock_minimo = db.Column(db.Float, default=0)
    precio_unitario = db.Column(db.Float, default=0)

class Venta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    total = db.Column(db.Float, nullable=False)
    metodo_pago = db.Column(db.String(20), nullable=False)  # efectivo, yape, tarjeta
    tipo_venta = db.Column(db.String(20), nullable=False)  # mostrador, delivery, takeout
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    estado = db.Column(db.String(50), default='pendiente') # pendiente, en_preparacion, listo, completado, cancelado
    estado_actualizado_en = db.Column(db.DateTime, nullable=True, default=datetime.utcnow, onupdate=datetime.utcnow)
    detalles = db.relationship('DetalleVenta', backref='venta', lazy=True, cascade="all, delete-orphan")
    delivery = db.relationship('Delivery', backref='venta', uselist=False, cascade="all, delete-orphan")
    usuario = db.relationship('Usuario', backref='ventas')

class DetalleVenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venta_id = db.Column(db.Integer, db.ForeignKey('venta.id'))
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'))
    cantidad = db.Column(db.Integer, nullable=False)
    precio_unitario = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    producto = db.relationship('Producto', backref='detalles_venta')

class Caja(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha_apertura = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_cierre = db.Column(db.DateTime)
    monto_inicial = db.Column(db.Float, default=0)
    monto_final = db.Column(db.Float)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    estado = db.Column(db.String(20), default='abierta')  # abierta, cerrada
    usuario = db.relationship('Usuario', backref='cajas')

class Egreso(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    concepto = db.Column(db.String(200), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    categoria = db.Column(db.String(50), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    usuario = db.relationship('Usuario', backref=db.backref('egresos', lazy=True))

class Compra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    proveedor = db.Column(db.String(100), nullable=False)
    total = db.Column(db.Float, nullable=False)
    factura = db.Column(db.String(50))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    detalles = db.relationship('DetalleCompra', backref='compra', lazy=True, cascade="all, delete-orphan")
    usuario = db.relationship('Usuario', backref='compras')

class DetalleCompra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    compra_id = db.Column(db.Integer, db.ForeignKey('compra.id'), nullable=False)
    insumo_id = db.Column(db.Integer, db.ForeignKey('insumo.id'), nullable=False)
    cantidad = db.Column(db.Float, nullable=False)
    precio_unitario = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    insumo = db.relationship('Insumo', backref='detalles_compra')

class MovimientoInsumo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    insumo_id = db.Column(db.Integer, db.ForeignKey('insumo.id'))
    tipo = db.Column(db.String(20), nullable=False)  # entrada, salida, inicial
    cantidad = db.Column(db.Float, nullable=False)
    concepto = db.Column(db.String(200), nullable=False)
    referencia_id = db.Column(db.Integer)  # ID de venta, compra, etc.
    stock_resultante = db.Column(db.Float, nullable=False)
    insumo = db.relationship('Insumo', backref='movimientos')

class Delivery(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venta_id = db.Column(db.Integer, db.ForeignKey('venta.id'))
    cliente_nombre = db.Column(db.String(100), nullable=False)
    direccion = db.Column(db.String(200), nullable=False)
    telefono = db.Column(db.String(20), nullable=False)
    hora_pedido = db.Column(db.DateTime, default=datetime.utcnow)
    hora_entrega_estimada = db.Column(db.DateTime)
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, en_camino, entregado

class Configuracion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(50), unique=True, nullable=False)
    valor = db.Column(db.String(200), nullable=False)

class CierreCaja(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha_cierre = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    monto_inicial = db.Column(db.Float, nullable=False)
    ventas_efectivo = db.Column(db.Float, nullable=False)
    ventas_tarjeta = db.Column(db.Float, nullable=False)
    total_ventas = db.Column(db.Float, nullable=False)
    total_egresos = db.Column(db.Float, nullable=False)
    monto_final_esperado = db.Column(db.Float, nullable=False)
    monto_final_real = db.Column(db.Float, nullable=False)
    diferencia = db.Column(db.Float, nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    usuario = db.relationship('Usuario', backref=db.backref('cierres_caja', lazy=True))

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Rutas principales
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = Usuario.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# Módulo de Ventas
@app.route('/ventas')
@login_required
def ventas():
    productos = Producto.query.filter_by(activo=True).all()
    return render_template('ventas.html', productos=productos)

@app.route('/api/crear_venta', methods=['POST'])
@login_required
def crear_venta():
    data = request.json
    venta = Venta(
        total=data['total'],
        metodo_pago=data['metodo_pago'],
        tipo_venta=data['tipo_venta'],
        usuario_id=current_user.id
    )
    db.session.add(venta)
    db.session.flush()
    
    for item in data['items']:
        detalle = DetalleVenta(
            venta_id=venta.id,
            producto_id=item['producto_id'],
            cantidad=item['cantidad'],
            precio_unitario=item['precio'],
            subtotal=item['subtotal']
        )
        db.session.add(detalle)

    # Si es delivery, crear el registro de delivery
    if data.get('tipo_venta') == 'delivery' and 'delivery_info' in data:
        info = data['delivery_info']
        if info.get('cliente_nombre') and info.get('direccion') and info.get('telefono'):
            delivery = Delivery(
                venta_id=venta.id,
                cliente_nombre=info['cliente_nombre'],
                direccion=info['direccion'],
                telefono=info['telefono'],
                hora_entrega_estimada=datetime.fromisoformat(info['hora_entrega'])
            )
            db.session.add(delivery)
    
    db.session.commit()
    return jsonify({'success': True, 'venta_id': venta.id})

@app.route('/api/generar_ticket/<int:venta_id>')
@login_required
def generar_ticket(venta_id):
    from reportlab.lib.pagesizes import landscape
    venta = Venta.query.get_or_404(venta_id)
    detalles = DetalleVenta.query.filter_by(venta_id=venta_id).all()
    buffer = io.BytesIO()
    # Tamaño tipo ticket: 220px ~ 6.2cm de ancho (rollo 58mm)
    ticket_width = 220
    ticket_height = 400 + 20 * len(detalles)
    from reportlab.pdfgen.canvas import Canvas
    p = Canvas(buffer, pagesize=(ticket_width, ticket_height))
    y = ticket_height - 20
    # Encabezado centrado
    p.setFont("Courier-Bold", 12)
    p.drawCentredString(ticket_width/2, y, "POLLERÍA JT")
    y -= 18
    p.setFont("Courier", 9)
    p.drawCentredString(ticket_width/2, y, f"Ticket #{venta_id}")
    y -= 13
    p.drawCentredString(ticket_width/2, y, f"Fecha: {venta.fecha.strftime('%d/%m/%Y %H:%M')}")
    y -= 13
    p.drawCentredString(ticket_width/2, y, f"Cajero: {current_user.nombre}")
    y -= 13
    # Separador
    p.setDash(2,2)
    p.line(10, y, ticket_width-10, y)
    p.setDash()
    y -= 10
    # Encabezados de tabla
    p.setFont("Courier-Bold", 8)
    p.drawString(10, y, "Producto")
    p.drawRightString(110, y, "Cant")
    p.drawRightString(150, y, "P.U.")
    p.drawRightString(ticket_width-10, y, "Subt.")
    y -= 10
    p.setFont("Courier", 8)
    for detalle in detalles:
        producto = Producto.query.get(detalle.producto_id)
        p.drawString(10, y, producto.nombre[:15])
        p.drawRightString(110, y, str(detalle.cantidad))
        p.drawRightString(150, y, f"S/. {detalle.precio_unitario:.2f}")
        p.drawRightString(ticket_width-10, y, f"S/. {detalle.subtotal:.2f}")
        y -= 10
    # Separador
    y -= 2
    p.setDash(2,2)
    p.line(10, y, ticket_width-10, y)
    p.setDash()
    y -= 12
    # Total
    p.setFont("Courier-Bold", 10)
    p.drawRightString(150, y, "TOTAL:")
    p.drawRightString(ticket_width-10, y, f"S/. {venta.total:.2f}")
    y -= 18
    # Mensaje de agradecimiento
    p.setFont("Courier-Oblique", 9)
    p.drawCentredString(ticket_width/2, y, "¡Gracias por su compra!")
    p.showPage()
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'ticket_{venta_id}.pdf')

# Módulo de Caja
@app.route('/caja')
@login_required
def caja():
    # Esta vista ahora solo renderiza la plantilla.
    # Todos los datos se cargarán dinámicamente vía API.
    return render_template('caja.html')

@app.route('/api/abrir_caja', methods=['POST'])
@login_required
def abrir_caja():
    """Endpoint para abrir una nueva caja."""
    data = request.get_json()
    if Caja.query.filter_by(estado='abierta').first():
        return jsonify({'success': False, 'message': 'Ya existe una caja abierta.'}), 400

    nueva_caja = Caja(
        monto_inicial=float(data.get('monto_inicial', 0)),
        usuario_id=current_user.id,
        fecha_apertura=datetime.utcnow() # Siempre UTC al crear
    )
    db.session.add(nueva_caja)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Caja abierta correctamente.'})

# Módulo de Inventario
@app.route('/inventario')
@login_required
def inventario():
    insumos = Insumo.query.order_by(Insumo.nombre).all()
    return render_template('inventario.html', insumos=insumos)

def insumo_to_dict(insumo):
    return {
        'id': insumo.id,
        'nombre': insumo.nombre,
        'unidad': insumo.unidad,
        'stock_actual': insumo.stock_actual,
        'stock_minimo': insumo.stock_minimo,
        'precio_unitario': insumo.precio_unitario
    }

@app.route('/api/agregar-insumo', methods=['POST'])
@login_required
def agregar_insumo_api():
    data = request.json
    
    if Insumo.query.filter_by(nombre=data['nombre']).first():
        return jsonify({'success': False, 'message': f"El insumo '{data['nombre']}' ya existe."})

    nuevo_insumo = Insumo(
        nombre=data['nombre'],
        unidad=data['unidad'],
        stock_actual=data['stock_actual'],
        stock_minimo=data['stock_minimo'],
        precio_unitario=data['precio_unitario']
    )
    db.session.add(nuevo_insumo)
    db.session.commit()

    if nuevo_insumo.stock_actual > 0:
        movimiento = MovimientoInsumo(
            insumo_id=nuevo_insumo.id,
            tipo='inicial',
            cantidad=nuevo_insumo.stock_actual,
            concepto='Stock inicial',
            stock_resultante=nuevo_insumo.stock_actual
        )
        db.session.add(movimiento)
        db.session.commit()

    return jsonify({'success': True, 'insumo': insumo_to_dict(nuevo_insumo)})

@app.route('/api/actualizar_stock', methods=['POST'])
@login_required
def actualizar_stock():
    data = request.json
    insumo = Insumo.query.get(data['insumo_id'])
    stock_anterior = insumo.stock_actual
    nuevo_stock = float(data['nuevo_stock'])
    
    insumo.stock_actual = nuevo_stock

    movimiento = MovimientoInsumo(
        insumo_id=insumo.id,
        tipo='salida' if nuevo_stock < stock_anterior else 'entrada',
        cantidad=abs(nuevo_stock - stock_anterior),
        concepto=data['concepto'],
        stock_resultante=nuevo_stock
    )
    db.session.add(movimiento)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/alertas-stock')
@login_required
def get_alertas_stock():
    alertas = Insumo.query.filter(Insumo.stock_actual <= Insumo.stock_minimo).all()
    return jsonify([{
        'id': a.id,
        'nombre': a.nombre,
        'stock_actual': a.stock_actual,
        'stock_minimo': a.stock_minimo,
        'unidad': a.unidad
    } for a in alertas])

@app.route('/api/movimientos-insumo/<int:insumo_id>')
@login_required
def get_movimientos(insumo_id):
    movimientos = MovimientoInsumo.query.filter_by(insumo_id=insumo_id).order_by(MovimientoInsumo.fecha.desc()).all()
    insumo = Insumo.query.get(insumo_id)
    if not insumo:
        return jsonify([]) # O un error 404
        
    return jsonify([{
        'fecha': m.fecha.strftime('%d/%m/%Y %H:%M'),
        'tipo': m.tipo,
        'cantidad': m.cantidad,
        'concepto': m.concepto,
        'stock_resultante': m.stock_resultante,
        'unidad': insumo.unidad
    } for m in movimientos])

# Módulo de Compras
@app.route('/compras')
@login_required
def compras():
    compras = Compra.query.order_by(Compra.fecha.desc()).all()
    insumos = Insumo.query.all()
    return render_template('compras.html', compras=compras, insumos=insumos)

@app.route('/api/registrar_compra', methods=['POST'])
@login_required
def registrar_compra():
    try:
        data = request.json
        
        # Crear la compra principal
        nueva_compra = Compra(
            proveedor=data['proveedor'],
            factura=data.get('factura'),
            total=data['total'],
            usuario_id=current_user.id  # <-- AQUÍ ESTÁ LA CORRECCIÓN
        )
        db.session.add(nueva_compra)
        db.session.flush()

        # Crear los detalles y actualizar stock
        for item in data['detalles']:
            insumo = Insumo.query.get(item['insumo_id'])
            if not insumo:
                db.session.rollback()
                return jsonify({'success': False, 'message': f"Insumo con ID {item['insumo_id']} no encontrado."}), 404

            detalle = DetalleCompra(
                compra_id=nueva_compra.id,
                insumo_id=item['insumo_id'],
                cantidad=item['cantidad'],
                precio_unitario=item['precio_unitario'],
                subtotal=item['subtotal']
            )
            db.session.add(detalle)
            
            # Actualizar stock y crear movimiento
            insumo.stock_actual += float(item['cantidad'])
            movimiento = MovimientoInsumo(
                insumo_id=insumo.id,
                tipo='entrada',
                cantidad=item['cantidad'],
                concepto=f"Compra a {data['proveedor']} (Fact: {data.get('factura', 'N/A')})",
                referencia_id=nueva_compra.id,
                stock_resultante=insumo.stock_actual
            )
            db.session.add(movimiento)

        db.session.commit()
        
        # Devolver datos para la actualización dinámica
        compra_data = {
            'id': nueva_compra.id,
            'fecha': nueva_compra.fecha.isoformat(),
            'proveedor': nueva_compra.proveedor,
            'factura': nueva_compra.factura,
            'total': nueva_compra.total,
            'usuario_nombre': current_user.nombre
        }
        
        return jsonify({'success': True, 'compra': compra_data})

    except Exception as e:
        db.session.rollback()
        # Log del error para depuración
        print(f"Error al registrar compra: {e}")
        return jsonify({'success': False, 'message': 'Ocurrió un error inesperado en el servidor.'}), 500

# Módulo de Cocina
@app.route('/cocina')
@login_required
def cocina():
    ventas_pendientes = Venta.query.filter_by(estado='pendiente').order_by(Venta.fecha).all()
    ventas_preparacion = Venta.query.filter_by(estado='en_preparacion').order_by(Venta.fecha).all()
    return render_template('cocina.html', 
                         ventas_pendientes=ventas_pendientes,
                         ventas_preparacion=ventas_preparacion)

@app.route('/api/cocina/pedidos')
@login_required
def api_cocina_pedidos():
    pendientes = Venta.query.filter_by(estado='pendiente').order_by(Venta.fecha.asc()).all()
    en_preparacion = Venta.query.filter_by(estado='en_preparacion').order_by(Venta.fecha.asc()).all()
    
    peru_tz = pytz.timezone('America/Lima')
    now_peru = datetime.now(peru_tz)
    start_of_day_peru = peru_tz.localize(datetime.combine(now_peru.date(), time.min))
    start_of_day_utc = start_of_day_peru.astimezone(pytz.utc)

    listos_hoy = Venta.query.filter(
        Venta.estado == 'listo',
        Venta.estado_actualizado_en != None,
        Venta.estado_actualizado_en >= start_of_day_utc
    ).count()
    
    completados = Venta.query.filter_by(estado='listo').order_by(Venta.estado_actualizado_en.desc()).limit(10).all()

    def Venta_a_dict(venta):
        # Convertir datetimes a formato ISO 8601 con información de zona horaria (UTC)
        fecha_utc_aware = venta.fecha.replace(tzinfo=pytz.utc) if venta.fecha else None
        inicio_prep_utc_aware = venta.estado_actualizado_en.replace(tzinfo=pytz.utc) if venta.estado_actualizado_en else None

        return {
            'id': venta.id,
            'total': venta.total,
            'tipo_venta': venta.tipo_venta,
            'estado': venta.estado,
            'fecha_utc': fecha_utc_aware.isoformat() if fecha_utc_aware else None,
            'detalles': [{'cantidad': d.cantidad, 'nombre_producto': d.producto.nombre, 'categoria': d.producto.categoria} for d in venta.detalles],
            'inicio_preparacion_utc': inicio_prep_utc_aware.isoformat() if inicio_prep_utc_aware and venta.estado == 'en_preparacion' else None
        }

    def format_completado(v):
         tiempo_preparacion = (v.estado_actualizado_en - v.fecha).total_seconds() / 60 if v.estado_actualizado_en else 0
         return {
            'id': v.id,
            'productos': ', '.join([f"{d.cantidad}x {d.producto.nombre}" for d in v.detalles]),
            'tiempo_preparacion': f"{tiempo_preparacion:.0f} min"
        }

    return jsonify({
        'pendientes': [Venta_a_dict(v) for v in pendientes],
        'en_preparacion': [Venta_a_dict(v) for v in en_preparacion],
        'completados': [format_completado(v) for v in completados],
        'stats': {
            'pendientes': len(pendientes),
            'en_preparacion': len(en_preparacion),
            'listos_hoy': listos_hoy
        }
    })

@app.route('/api/actualizar_estado_venta', methods=['POST'])
@login_required
def actualizar_estado_venta():
    data = request.json
    venta_id = data.get('venta_id')
    nuevo_estado = data.get('nuevo_estado')
    
    venta = Venta.query.get(venta_id)
    if not venta:
        return jsonify({'status': 'error', 'message': 'Venta no encontrada'}), 404
        
    venta.estado = nuevo_estado
    venta.estado_actualizado_en = datetime.utcnow() # Aseguramos la actualización de la hora
    db.session.commit()
    
    return jsonify({'status': 'success', 'message': 'Estado actualizado correctamente'})

@app.route('/api/pedidos-completados')
@login_required
def pedidos_completados():
    completados = Venta.query.filter_by(estado='listo').order_by(Venta.estado_actualizado_en.desc()).all()
    
    response_data = []
    for v in completados:
        # Calcular tiempo de preparación
        tiempo_preparacion = "N/A"
        if v.fecha_actualizacion_estado and v.fecha:
            diff = v.fecha_actualizacion_estado - v.fecha
            minutes = diff.total_seconds() / 60
            tiempo_preparacion = f"{int(minutes)} min"

        response_data.append({
            'id': v.id,
            'fecha': v.fecha.strftime('%H:%M'),
            'total': f"S/. {v.total:.2f}",
            'estado': v.estado,
            'tiempo_preparacion': tiempo_preparacion,
            'tiempo_promedio': '25 min'
        })

    return jsonify(response_data)

# Módulo de Delivery
@app.route('/delivery')
@login_required
def delivery():
    # Consultar todos los deliveries con su información de venta asociada
    deliveries = Delivery.query.join(Venta).order_by(Delivery.hora_pedido.desc()).all()
    
    # Separar los deliveries por estado
    pendientes = [d for d in deliveries if d.estado == 'pendiente']
    en_camino = [d for d in deliveries if d.estado == 'en_camino']
    
    # Para el historial, podemos mostrar todos o solo los entregados
    historial = deliveries
    
    return render_template('delivery.html', 
                           pendientes=pendientes, 
                           en_camino=en_camino, 
                           historial=historial)

@app.route('/api/registrar_delivery', methods=['POST'])
@login_required
def registrar_delivery():
    data = request.json
    delivery = Delivery(
        venta_id=data['venta_id'],
        cliente_nombre=data['cliente_nombre'],
        direccion=data['direccion'],
        telefono=data['telefono'],
        hora_entrega_estimada=datetime.fromisoformat(data['hora_entrega_estimada'])
    )
    db.session.add(delivery)
    db.session.commit()
    return jsonify({'success': True})

# Módulo de Reportes
@app.route('/reportes')
@login_required
def reportes():
    return render_template('reportes.html')

@app.route('/api/reporte_ventas')
@login_required
def reporte_ventas():
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    query = Venta.query
    if fecha_inicio:
        query = query.filter(Venta.fecha >= datetime.fromisoformat(fecha_inicio))
    if fecha_fin:
        query = query.filter(Venta.fecha <= datetime.fromisoformat(fecha_fin))
    
    ventas = query.all()
    
    # Agrupar por día
    ventas_por_dia = {}
    for venta in ventas:
        fecha = venta.fecha.strftime('%Y-%m-%d')
        if fecha not in ventas_por_dia:
            ventas_por_dia[fecha] = {'total': 0, 'cantidad': 0}
        ventas_por_dia[fecha]['total'] += venta.total
        ventas_por_dia[fecha]['cantidad'] += 1
    
    return jsonify(ventas_por_dia)

@app.route('/api/deliveries')
@login_required
def get_deliveries():
    # Obtenemos todos los deliveries para las diferentes secciones
    deliveries_pendientes = Delivery.query.filter_by(estado='pendiente').order_by(Delivery.hora_pedido.asc()).all()
    deliveries_en_camino = Delivery.query.filter_by(estado='en_camino').order_by(Delivery.hora_pedido.asc()).all()
    # Historial podría incluir entregados y cancelados del día, por ejemplo
    start_of_day = datetime.combine(date.today(), time.min)
    deliveries_historial = Delivery.query.filter(Delivery.estado.in_(['entregado', 'cancelado'])).filter(Delivery.hora_pedido >= start_of_day).order_by(Delivery.hora_pedido.desc()).all()

    # Calcular estadísticas
    entregados_hoy = Delivery.query.filter_by(estado='entregado').filter(Delivery.hora_pedido >= start_of_day).count()
    
    # Formatear datos para la respuesta JSON
    def format_delivery(d):
        return {
            'id': d.id,
            'venta_id': d.venta_id,
            'cliente_nombre': d.cliente_nombre,
            'direccion': d.direccion,
            'telefono': d.telefono,
            'hora_pedido_iso': d.hora_pedido.isoformat(),
            'hora_entrega_estimada_iso': d.hora_entrega_estimada.isoformat() if d.hora_entrega_estimada else None,
            'estado': d.estado
        }

    data = {
        'pendientes': [format_delivery(d) for d in deliveries_pendientes],
        'en_camino': [format_delivery(d) for d in deliveries_en_camino],
        'historial': [format_delivery(d) for d in deliveries_historial],
        'stats': {
            'pendientes': len(deliveries_pendientes),
            'en_camino': len(deliveries_en_camino),
            'entregados_hoy': entregados_hoy,
            'tiempo_promedio': '28 min' # Este valor podría ser calculado dinámicamente
        }
    }
    return jsonify(data)

@app.route('/api/delivery/<int:delivery_id>/estado', methods=['POST'])
@login_required
def actualizar_estado_delivery(delivery_id):
    data = request.get_json()
    nuevo_estado = data.get('estado')

    delivery = Delivery.query.get_or_404(delivery_id)
    
    if nuevo_estado not in ['en_camino', 'entregado', 'cancelado']:
        return jsonify({'success': False, 'message': 'Estado no válido'}), 400

    delivery.estado = nuevo_estado
    if nuevo_estado == 'entregado':
        delivery.hora_entrega_real = datetime.utcnow()

    db.session.commit()
    return jsonify({'success': True, 'message': 'Estado actualizado correctamente'})

@app.route('/api/detalle-venta/<int:venta_id>')
@login_required
def detalle_venta(venta_id):
    venta = Venta.query.get_or_404(venta_id)
    detalles = DetalleVenta.query.filter_by(venta_id=venta_id).all()
    
    response_data = {
        'id': venta.id,
        'fecha': venta.fecha.strftime('%d/%m/%Y %H:%M'),
        'total': f"S/. {venta.total:.2f}",
        'metodo_pago': venta.metodo_pago,
        'tipo_venta': venta.tipo_venta,
        'estado': venta.estado,
        'tiempo_preparacion': venta.fecha_actualizacion_estado.strftime('%H:%M') if venta.fecha_actualizacion_estado else "N/A"
    }
    
    detalles_data = []
    for detalle in detalles:
        producto = Producto.query.get(detalle.producto_id)
        detalles_data.append({
            'id': detalle.id,
            'producto': producto.nombre,
            'cantidad': detalle.cantidad,
            'precio_unitario': f"S/. {detalle.precio_unitario:.2f}",
            'subtotal': f"S/. {detalle.subtotal:.2f}"
        })
    
    response_data['detalles'] = detalles_data
    
    return jsonify(response_data)

@app.route('/api/detalle-delivery/<int:delivery_id>')
@login_required
def detalle_delivery(delivery_id):
    delivery = Delivery.query.get_or_404(delivery_id)
    response_data = {
        'id': delivery.id,
        'venta_id': delivery.venta_id,
        'cliente_nombre': delivery.cliente_nombre,
        'direccion': delivery.direccion,
        'telefono': delivery.telefono,
        'hora_pedido': delivery.hora_pedido.strftime('%d/%m/%Y %H:%M'),
        'hora_entrega_estimada': delivery.hora_entrega_estimada.strftime('%d/%m/%Y %H:%M') if delivery.hora_entrega_estimada else "N/A",
        'estado': delivery.estado
    }
    return jsonify(response_data)

@app.route('/api/reporte')
@login_required
def generar_reporte():
    tipo_reporte = request.args.get('tipo', 'ventas')
    fecha_inicio_str = request.args.get('fecha_inicio')
    fecha_fin_str = request.args.get('fecha_fin')

    # Conversión de fechas robusta
    try:
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        else:
            fecha_inicio = datetime.utcnow()
        
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        else:
            fecha_fin = datetime.utcnow()

        # Asegurar que el rango de fechas cubra los días completos
        fecha_inicio = datetime.combine(fecha_inicio.date(), time.min)
        fecha_fin = datetime.combine(fecha_fin.date(), time.max)
    except ValueError:
        return jsonify({'success': False, 'message': 'Formato de fecha inválido. Usar YYYY-MM-DD.'}), 400

    # Inicializar respuesta
    respuesta = {
        'stats': {},
        'charts': {
            'categorias': {'labels': [], 'data': []},
            'ventas': {'labels': [], 'data': []}
        },
        'table_data': []
    }

    if tipo_reporte == 'ventas':
        # Consulta principal de ventas en el rango
        ventas_query = Venta.query.filter(Venta.fecha.between(fecha_inicio, fecha_fin))
        
        # 1. Stats
        ventas_list = ventas_query.all()
        total_ventas = sum(v.total for v in ventas_list)
        total_pedidos = len(ventas_list)
        respuesta['stats'] = {
            'ventas_totales': total_ventas,
            'total_pedidos': total_pedidos,
            'promedio_pedido': total_ventas / total_pedidos if total_pedidos > 0 else 0,
            'total_deliveries': ventas_query.filter_by(tipo_venta='delivery').count()
        }

        # 2. Chart - Ventas por categoría
        ventas_por_categoria = db.session.query(
            Producto.categoria, func.sum(DetalleVenta.subtotal)
        ).join(DetalleVenta, DetalleVenta.producto_id == Producto.id)\
         .join(Venta, Venta.id == DetalleVenta.venta_id)\
         .filter(Venta.fecha.between(fecha_inicio, fecha_fin))\
         .group_by(Producto.categoria).all()
        
        if ventas_por_categoria:
            respuesta['charts']['categorias']['labels'] = [vc[0] for vc in ventas_por_categoria]
            respuesta['charts']['categorias']['data'] = [float(vc[1]) for vc in ventas_por_categoria]

        # 3. Chart - Ventas por día
        ventas_por_dia = db.session.query(
            func.date(Venta.fecha), func.sum(Venta.total)
        ).filter(Venta.fecha.between(fecha_inicio, fecha_fin))\
         .group_by(func.date(Venta.fecha)).order_by(func.date(Venta.fecha)).all()
        
        if ventas_por_dia:
            labels = []
            data = []
            for row in ventas_por_dia:
                fecha_obj = row[0]
                # Si es un string, lo convertimos a objeto datetime. Si ya es un objeto date, funciona igual.
                if isinstance(fecha_obj, str):
                    try:
                        fecha_obj = datetime.strptime(fecha_obj, '%Y-%m-%d').date()
                    except ValueError:
                        continue # O manejar el error como prefieras
                
                labels.append(fecha_obj.strftime('%d/%m'))
                data.append(float(row[1]))

            respuesta['charts']['ventas']['labels'] = labels
            respuesta['charts']['ventas']['data'] = data

        # 4. Table Data
        respuesta['table_data'] = [{
            'id': v.id,
            # Convertir a hora de Perú antes de mostrar
            'fecha': datetime.strptime(v.fecha.strftime('%Y-%m-%d %H:%M:%S'), '%Y-%m-%d %H:%M:%S').replace(tzinfo=pytz.utc).astimezone(pytz.timezone('America/Lima')).strftime('%d/%m/%Y %H:%M'),
            'productos': ", ".join([d.producto.nombre for d in v.detalles]),
            'tipo': v.tipo_venta,
            'metodo_pago': v.metodo_pago,
            'total': v.total
        } for v in ventas_list]

    elif tipo_reporte == 'stock_bajo':
        insumos_bajos = Insumo.query.filter(Insumo.stock_actual <= Insumo.stock_minimo).all()
        respuesta['table_data'] = [{
            'nombre': i.nombre,
            'stock_actual': i.stock_actual,
            'stock_minimo': i.stock_minimo,
            'unidad': i.unidad
        } for i in insumos_bajos]

    return jsonify(respuesta)

@app.route('/api/exportar-excel')
@login_required
def exportar_excel():
    tipo = request.args.get('tipo', 'ventas')
    fecha_inicio_str = request.args.get('fecha_inicio')
    fecha_fin_str = request.args.get('fecha_fin')

    if not fecha_inicio_str or not fecha_fin_str:
        return "Fechas no proporcionadas", 400

    fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
    fecha_fin = datetime.combine(datetime.strptime(fecha_fin_str, '%Y-%m-%d').date(), time.max)

    ventas = Venta.query.filter(Venta.fecha.between(fecha_inicio, fecha_fin)).order_by(Venta.fecha.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # --- Estilos ---
    header_font = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    currency_format = 'S/ #,##0.00'
    date_format = 'DD/MM/YYYY HH:MM'
    
    # --- Título ---
    ws.merge_cells('A1:F2')
    title_cell = ws['A1']
    title_cell.value = f"Reporte de Ventas del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    title_cell.font = Font(name='Calibri', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Encabezados ---
    headers = ["Fecha", "ID Venta", "Productos", "Tipo", "Método de Pago", "Total"]
    ws.append(headers) # Se insertan en la fila 3
    
    header_row = ws[3]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill

    # --- Datos y Subtotales por Día ---
    ventas_por_dia = defaultdict(list)
    for v in ventas:
        productos_str = ", ".join([d.producto.nombre for d in v.detalles])
        fecha_utc = v.fecha.replace(tzinfo=pytz.utc)
        peru_tz = pytz.timezone('America/Lima')
        fecha_peru = fecha_utc.astimezone(peru_tz)
        fecha_peru_naive = fecha_peru.replace(tzinfo=None)
        fecha_dia = fecha_peru_naive.date()
        ventas_por_dia[fecha_dia].append((fecha_peru_naive, v.id, productos_str, v.tipo_venta, v.metodo_pago, v.total))

    fila_actual = 4
    total_general = 0
    for dia, ventas_dia in sorted(ventas_por_dia.items()):
        subtotal_dia = 0
        for row in ventas_dia:
            ws.append(row)
            subtotal_dia += row[5]
            fila_actual += 1
        # Fila de subtotal
        ws.append(["", "", "", "", "Subtotal Día", subtotal_dia])
        ws[f'F{fila_actual+1}'].font = Font(bold=True, color='006100')
        ws[f'E{fila_actual+1}'].font = Font(bold=True, color='006100')
        ws[f'E{fila_actual+1}'].alignment = Alignment(horizontal='right')
        ws[f'F{fila_actual+1}'].number_format = currency_format
        fila_actual += 1
        total_general += subtotal_dia
    # Fila de total general
    ws.append(["", "", "", "", "TOTAL GENERAL", total_general])
    ws[f'F{fila_actual+1}'].font = Font(bold=True, color='C00000')
    ws[f'E{fila_actual+1}'].font = Font(bold=True, color='C00000')
    ws[f'E{fila_actual+1}'].alignment = Alignment(horizontal='right')
    ws[f'F{fila_actual+1}'].number_format = currency_format

    # --- Formatos y Ancho de Columnas ---
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = col[2].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        if column == 'A':
             ws.column_dimensions[column].width = 20
             for cell in col:
                 if isinstance(cell.value, datetime):
                    cell.number_format = date_format
        if column == 'F':
            for cell in col:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = currency_format

    ws.freeze_panes = 'A4'

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    return send_file(
        virtual_workbook,
        as_attachment=True,
        download_name=f'reporte_ventas_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/configurar-exportacion', methods=['POST'])
@login_required
def configurar_exportacion():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400

    try:
        # Guardar si la exportación automática está habilitada
        auto_export_setting = Configuracion.query.filter_by(clave='auto_export_enabled').first()
        if not auto_export_setting:
            auto_export_setting = Configuracion(clave='auto_export_enabled', valor=str(data.get('auto_export', False)))
            db.session.add(auto_export_setting)
        else:
            auto_export_setting.valor = str(data.get('auto_export', False))

        # Guardar la frecuencia
        frecuencia_setting = Configuracion.query.filter_by(clave='auto_export_frequency').first()
        if not frecuencia_setting:
            frecuencia_setting = Configuracion(clave='auto_export_frequency', valor=data.get('frecuencia', 'diario'))
            db.session.add(frecuencia_setting)
        else:
            frecuencia_setting.valor = data.get('frecuencia', 'diario')

        db.session.commit()
        return jsonify({'success': True, 'message': 'Configuración guardada correctamente'})
    except Exception as e:
        db.session.rollback()
        # En una aplicación real, se registraría el error: app.logger.error(f"Error al guardar config: {e}")
        return jsonify({'success': False, 'message': 'Error interno al guardar la configuración'}), 500

@app.route('/api/exportar-pdf')
@login_required
def exportar_pdf():
    # Lógica similar para PDF con reportlab
    return "Función de exportar a PDF no implementada", 501

@app.route('/api/caja/datos')
@login_required
def get_caja_datos():
    caja_actual = Caja.query.filter_by(estado='abierta').first()
    
    caja_info = None
    total_ventas = 0
    total_egresos = 0
    egresos_del_turno = []

    if caja_actual:
        ventas_del_turno = Venta.query.filter(Venta.fecha >= caja_actual.fecha_apertura).all()
        egresos_del_turno = Egreso.query.filter(Egreso.fecha >= caja_actual.fecha_apertura).all()

        total_ventas = sum(v.total for v in ventas_del_turno)
        total_egresos = sum(e.monto for e in egresos_del_turno)

        # CONVERSIÓN DE ZONA HORARIA
        utc_time = caja_actual.fecha_apertura.replace(tzinfo=pytz.utc)
        peru_tz = pytz.timezone('America/Lima')
        peru_time = utc_time.astimezone(peru_tz)

        caja_info = {
            'id': caja_actual.id,
            'monto_inicial': caja_actual.monto_inicial,
            'usuario': caja_actual.usuario.nombre,
            'fecha_apertura': peru_time.strftime('%d/%m/%Y a las %H:%M')
        }
    
    def format_egreso(e):
        return {
            'id': e.id,
            'hora': e.fecha.strftime('%H:%M'),
            'concepto': e.concepto,
            'categoria': e.categoria,
            'monto': e.monto,
            'registrado_por': e.usuario.username
        }

    return jsonify({
        'caja_info': caja_info,
        'total_ventas': total_ventas,
        'total_egresos': total_egresos,
        'balance': (caja_info['monto_inicial'] if caja_info else 0) + total_ventas - total_egresos,
        'egresos': [format_egreso(e) for e in egresos_del_turno]
    })

@app.route('/api/caja/registrar-egreso', methods=['POST'])
@login_required
def registrar_egreso():
    data = request.get_json()
    if not data or not data.get('concepto') or not data.get('monto'):
        return jsonify({'success': False, 'message': 'Datos incompletos'}), 400
    
    nuevo_egreso = Egreso(
        concepto=data['concepto'],
        monto=float(data['monto']),
        categoria=data.get('categoria', 'Varios'),
        usuario_id=current_user.id
    )
    db.session.add(nuevo_egreso)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Egreso registrado correctamente'})

@app.route('/api/caja/cerrar', methods=['POST'])
@login_required
def cerrar_caja():
    """
    Endpoint simple para cerrar la caja.
    1. Calcula los totales del turno.
    2. Crea el registro histórico en CierreCaja.
    3. Actualiza el estado de la caja actual a 'cerrada'.
    """
    data = request.get_json()
    caja_actual = Caja.query.filter_by(estado='abierta').first()

    if not caja_actual:
        return jsonify({'success': False, 'message': 'No se encontró una caja abierta.'}), 400

    ventas_del_turno = Venta.query.filter(Venta.fecha >= caja_actual.fecha_apertura).all()
    egresos_del_turno = Egreso.query.filter(Egreso.fecha >= caja_actual.fecha_apertura).all()

    monto_inicial = caja_actual.monto_inicial
    ventas_efectivo = sum(v.total for v in ventas_del_turno if v.metodo_pago == 'efectivo')
    ventas_tarjeta = sum(v.total for v in ventas_del_turno if v.metodo_pago != 'efectivo')
    total_ventas = ventas_efectivo + ventas_tarjeta
    total_egresos = sum(e.monto for e in egresos_del_turno)
    
    monto_final_esperado = monto_inicial + ventas_efectivo - total_egresos
    monto_final_real = float(data.get('monto_final_real', 0))
    diferencia = monto_final_real - monto_final_esperado

    # 2. Crear el registro histórico
    cierre = CierreCaja(
        monto_inicial=monto_inicial,
        ventas_efectivo=ventas_efectivo,
        ventas_tarjeta=ventas_tarjeta,
        total_ventas=total_ventas,
        total_egresos=total_egresos,
        monto_final_esperado=monto_final_esperado,
        monto_final_real=monto_final_real,
        diferencia=diferencia,
        usuario_id=current_user.id,
        fecha_cierre=datetime.utcnow()
    )
    db.session.add(cierre)

    # 3. Actualizar la caja actual
    caja_actual.estado = 'cerrada'
    caja_actual.fecha_cierre = datetime.utcnow()
    caja_actual.monto_final = monto_final_real
    
    db.session.commit()

    return jsonify({'success': True, 'message': 'Caja cerrada exitosamente'})

@app.route('/api/pedidos-recientes')
@login_required
def pedidos_recientes():
    ventas = Venta.query.order_by(Venta.fecha.desc()).limit(10).all()
    pedidos = []
    for v in ventas:
        # Convertir fecha a hora de Perú
        fecha_utc = v.fecha.replace(tzinfo=pytz.utc)
        peru_tz = pytz.timezone('America/Lima')
        fecha_peru = fecha_utc.astimezone(peru_tz)
        pedidos.append({
            'id': v.id,
            'fecha': fecha_peru.strftime('%d/%m/%Y %H:%M'),
            'total': f"S/. {v.total:.2f}",
            'estado': v.estado
        })
    return jsonify(pedidos)

@app.route('/crear_usuario_test')
def crear_usuario_test():
    from werkzeug.security import generate_password_hash
    nuevo = Usuario(username='testuser', password_hash=generate_password_hash('test123'), nombre='Test User', rol='admin', activo=True)
    db.session.add(nuevo)
    db.session.commit()
    return 'Usuario testuser creado'

# MIGRACIÓN DE BASE DE DATOS SIMPLE
with app.app_context():
    inspector = inspect(db.engine)
    table_name = Venta.__tablename__
    if not any(c['name'] == 'estado_actualizado_en' for c in inspector.get_columns(table_name)):
        print(f"Añadiendo columna 'estado_actualizado_en' a la tabla '{table_name}'...")
        # Usamos una conexión para ejecutar el ALTER TABLE
        with db.engine.connect() as connection:
            connection.execute(text(f'ALTER TABLE {table_name} ADD COLUMN estado_actualizado_en DATETIME;'))
            connection.commit() # Asegurarse de que el cambio se guarde
            print("Columna añadida con éxito.")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Crear usuario admin por defecto si no existe
        if not Usuario.query.filter_by(username='admin').first():
            admin = Usuario(
                username='admin',
                password_hash=generate_password_hash('admin123'),
                nombre='Administrador',
                rol='admin'
            )
            db.session.add(admin)
            db.session.commit()
            
            # Crear productos de ejemplo si no existen
            if not Producto.query.first():
                productos_ejemplo = [
                    Producto(nombre='Combo Pollo a la Brasa', precio=25.00, categoria='combo'),
                    Producto(nombre='1/4 Pollo a la Brasa', precio=15.00, categoria='plato'),
                    Producto(nombre='1/2 Pollo a la Brasa', precio=28.00, categoria='plato'),
                    Producto(nombre='Pollo Entero', precio=50.00, categoria='plato'),
                    Producto(nombre='Papas Fritas', precio=8.00, categoria='plato'),
                    Producto(nombre='Ensalada César', precio=12.00, categoria='plato'),
                    Producto(nombre='Gaseosa', precio=5.00, categoria='bebida'),
                    Producto(nombre='Chicha Morada', precio=4.00, categoria='bebida'),
                    Producto(nombre='Flan', precio=6.00, categoria='postre')
                ]
                db.session.add_all(productos_ejemplo)
            
            # Crear insumos de ejemplo si no existen
            if not Insumo.query.first():
                insumos_ejemplo = [
                    Insumo(nombre='Pollo', unidad='kg', stock_actual=100, stock_minimo=20, precio_unitario=12.00),
                    Insumo(nombre='Papas', unidad='kg', stock_actual=50, stock_minimo=10, precio_unitario=3.50),
                    Insumo(nombre='Aceite', unidad='litros', stock_actual=20, stock_minimo=5, precio_unitario=8.00),
                    Insumo(nombre='Gaseosas', unidad='unidades', stock_actual=100, stock_minimo=20, precio_unitario=2.50)
                ]
                db.session.add_all(insumos_ejemplo)
            
            db.session.commit()
    
    app.run(debug=True) 